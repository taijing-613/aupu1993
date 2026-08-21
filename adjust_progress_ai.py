import openpyxl

p = r'C:\Users\J\WorkBuddy\2026-08-17-08-58-49\淘宝巡检系统_项目进度_20260821.xlsx'
wb = openpyxl.load_workbook(p)
ws = wb['模块进度总览']
# 第10行：AI 视觉比对 —— 已重构为 GPT-4V 语义级，待真实回归
ws.cell(10, 2, 'AI 视觉比对（多模态大模型 GPT-4V）')
ws.cell(10, 3, '60%')
ws.cell(10, 4, '进行中（待真实回归）')
ws.cell(10, 5,
        '已重构：弃用 SSIM/pHash 像素算法，接入 OpenAI GPT-4V 多模态大模型做语义级视觉合规判定'
        '（主图/SKU/详情三类），可识别价格篡改、违规文案、盗图等；模型不可用时自动回退像素比对并标注。'
        '待用户填入 API Key + 联网做真实召回回归验证。')
# 重算整体加权
t, n = 0, 0
for r in range(5, 19):
    v = ws.cell(r, 3).value
    if isinstance(v, str) and v.endswith('%'):
        t += float(v[:-1]); n += 1
avg = round(t / n)
ws.cell(19, 3, f'{avg}%')
ws.cell(19, 4, '核心四项未完全达标')
ws.cell(19, 5,
        '抓取40% / AI视觉比对60%(待真实回归) / 网站稳定性50% / 操作面板70%'
        ' —— 比对引擎已换真 AI，仍待 Key + 联网验证')
wb.save(p)
print('整体加权 =', avg, '%')
