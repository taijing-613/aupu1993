"""FastAPI 后端：真实 CRUD + 图片存盘 + 异步巡检（真实 SSIM/pHash/价格比对）。

核心升级：
1) 违规去重：每次巡检对同一 (shop_id, type) 只维护一条记录，重复巡检就地更新，绝不反复新增。
2) 一键巡店：仅需官旗 URL + 代理商 URL，后端用 Playwright 真实抓取主图/SKU/价格/详情页并比对；
   沙箱无网时自动降级为演示用图（明确标记 simulated），真实链路与降级链路共用同一比对引擎。
"""
import os
import io
import json
import time
from urllib.parse import urlparse
import uuid
import random
import threading
import asyncio
import datetime
import collections
import queue

import uvicorn
from fastapi import FastAPI, Request, UploadFile, File, HTTPException, Form, Response
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles

import db
import compare
import crawler
from PIL import Image, ImageDraw

BASE = os.path.dirname(os.path.abspath(__file__))
UPLOAD = db.UPLOAD
INDEX = os.path.join(BASE, "index.html")

db.init()
app = FastAPI(title="AI电商视觉合规巡检系统")
app.mount("/uploads", StaticFiles(directory=UPLOAD), name="uploads")


# ---------------- 实时协作：内存事件总线 + 在线状态 ----------------
# 单进程内广播（部署在单实例 PaaS 上即可满足小团队协作；多实例需换 Redis 发布订阅）。
_EVENTS = collections.deque()
_EVENT_LOCK = threading.Lock()
_EVENT_ID = [0]
_ONLINE = {}          # member_id -> 最后心跳时间戳（epoch）
_ONLINE_LOCK = threading.Lock()


def _publish(resource, action, actor_id, actor_name, detail, ref_id=None):
    """把一条事件推入总线，所有 SSE 客户端都会收到。"""
    _EVENT_ID[0] += 1
    ev = {
        "id": _EVENT_ID[0], "resource": resource, "action": action,
        "actor_id": actor_id, "actor_name": actor_name, "detail": detail,
        "ref_id": ref_id, "ts": time.strftime("%Y-%m-%d %H:%M:%S"),
    }
    with _EVENT_LOCK:
        _EVENTS.append(ev)
        if len(_EVENTS) > 300:
            _EVENTS.popleft()
    return ev


def _actor(request):
    """从 X-User-Id 头解析出当前操作者，找不到返回 (None, None)。"""
    if not request:
        return None, None
    uid = request.headers.get("X-User-Id")
    if uid:
        try:
            m = db.get_member(int(uid))
            if m:
                return m["id"], m["name"]
        except Exception:
            pass
    return None, None


def broadcast(resource, action, detail, request=None, ref_id=None,
              actor_id=None, actor_name=None):
    """写操作后调用：落库活动日志 + 广播实时事件。"""
    if actor_id is None or actor_name is None:
        actor_id, actor_name = _actor(request)
    db.log_activity(actor_id, actor_name, action, resource, detail, ref_id)
    _publish(resource, action, actor_id, actor_name, detail, ref_id)


def online_members():
    """返回 60 秒内心跳过的在线成员。"""
    now = time.time()
    with _ONLINE_LOCK:
        ids = [mid for mid, ts in _ONLINE.items() if now - ts < 60]
        for mid in list(_ONLINE):
            if now - _ONLINE[mid] >= 60:
                del _ONLINE[mid]
    if not ids:
        return []
    c = db.conn()
    rows = c.execute(
        "SELECT id,name,department FROM members WHERE id IN (%s)"
        % ",".join("?" * len(ids)), ids
    ).fetchall()
    c.close()
    return [dict(r) for r in rows]


def row_to_dict(r):
    return dict(r) if r else None


def rows_to_list(rs):
    return [dict(r) for r in rs]


def save_upload(file: UploadFile) -> str:
    ext = os.path.splitext(file.filename or "img")[1].lower() or ".png"
    name = uuid.uuid4().hex + ext
    path = os.path.join(UPLOAD, name)
    with open(path, "wb") as f:
        f.write(file.file.read())
    return name


# ---------------- 种子数据（生成可演示的真实比对图） ----------------
def _gen_image(path, has_logo=True, variant=0):
    img = Image.new("RGB", (600, 600), "white")
    d = ImageDraw.Draw(img)
    d.rectangle([0, 0, 600, 80], fill="#111827")
    d.text((20, 30), "OFFICIAL" if has_logo else "AGENT STORE", fill="white")
    if has_logo:
        d.rectangle([40, 120, 160, 200], outline="#4F46E5", width=5)
        d.text((52, 150), "LOGO", fill="#4F46E5")
    d.rectangle([40, 260, 560, 540], outline="#E4E4E7", width=2)
    if variant == 1:
        d.rectangle([40, 260, 560, 540], fill="#F4F4F5")
    img.save(path)


def _gen_synthetic(url: str, has_logo: bool = True) -> str:
    """沙箱降级：按 URL 确定性生成一张图，使不同 URL 得到不同的真实相似度分数。"""
    h = int(uuid.uuid5(uuid.NAMESPACE_URL, url).hex[:8], 16)
    rnd = random.Random(h)
    img = Image.new("RGB", (600, 600), "white")
    d = ImageDraw.Draw(img)
    d.rectangle([0, 0, 600, 80], fill="#111827")
    d.text((20, 30), "OFFICIAL" if has_logo else "AGENT", fill="white")
    if has_logo:
        d.rectangle([40, 120, 160, 200], outline="#4F46E5", width=5)
        d.text((52, 150), "LOGO", fill="#4F46E5")
    for _ in range(rnd.randint(3, 9)):
        x = rnd.randint(60, 520); y = rnd.randint(280, 520)
        w = rnd.randint(40, 120); hh = rnd.randint(40, 120)
        shade = rnd.randint(220, 245)
        d.rectangle([x, y, x + w, y + hh], fill=(shade, shade, shade))
    if not has_logo and rnd.random() < 0.5:
        d.rectangle([40, 120, 160, 200], outline="#9ca3af", width=3)
    name = uuid.uuid5(uuid.NAMESPACE_URL, url).hex[:12] + ".png"
    img.save(os.path.join(UPLOAD, name))
    return name


def seed_if_empty():
    c = db.conn()
    if c.execute("SELECT COUNT(*) FROM shops").fetchone()[0] > 0:
        c.close()
        return
    f = "seed_flag.png"
    _gen_image(os.path.join(UPLOAD, f), has_logo=True)
    c.execute(
        "INSERT INTO shops(name,owner,email,url,grp,role,price,image) VALUES(?,?,?,?,?,?,?,?)",
        ("官方旗舰店", "品牌部", "brand@corp.com", "https://flagship.example.com", "基准库", "flag", 199.0, f),
    )
    a1 = "seed_agent1.png"
    _gen_image(os.path.join(UPLOAD, a1), has_logo=False)
    c.execute(
        "INSERT INTO shops(name,owner,email,url,grp,role,price,image) VALUES(?,?,?,?,?,?,?,?)",
        ("代理商·潮流前线", "李 工", "li@agent.com", "https://agent-a.example.com", "渠道店", "agent", 159.0, a1),
    )
    a2 = "seed_agent2.png"
    _gen_image(os.path.join(UPLOAD, a2), has_logo=True, variant=1)
    c.execute(
        "INSERT INTO shops(name,owner,email,url,grp,role,price,image) VALUES(?,?,?,?,?,?,?,?)",
        ("代理商·优选购", "王 敏", "wang@agent.com", "https://agent-b.example.com", "渠道店", "agent", 209.0, a2),
    )
    s = "seed_std.png"
    _gen_image(os.path.join(UPLOAD, s), has_logo=True)
    c.execute(
        "INSERT INTO standards(name,page_type,hex_color,logo_x,logo_y,logo_w,logo_h,forbidden_fonts,image) VALUES(?,?,?,?,?,?,?,?,?)",
        ("主图基准-官旗", "主图", "#4F46E5", 6.7, 20.0, 20.0, 13.3, "楷体, 艺术字", s),
    )
    c.commit()
    c.close()


seed_if_empty()


def dedupe_inspections():
    """启动清理：同一对 (agent_url, official_url) 仅保留最新一条巡检记录。"""
    c = db.conn()
    try:
        rows = c.execute(
            "SELECT agent_url, official_url, GROUP_CONCAT(id) AS ids, COUNT(*) AS n "
            "FROM inspections GROUP BY agent_url, official_url HAVING n > 1"
        ).fetchall()
        for r in rows:
            keep = max(int(x) for x in r["ids"].split(","))
            for x in r["ids"].split(","):
                if int(x) != keep:
                    c.execute("DELETE FROM inspections WHERE id=?", (int(x),))
        c.commit()
    except Exception:
        pass
    c.close()


dedupe_inspections()


# ---------------- 通用 JSON 助手 ----------------
def jbody(request):
    return request.json()


@app.get("/api/health")
def health():
    return {"ok": True}


# ---------------- 实时协作端点 ----------------
@app.get("/api/events")
async def sse_events():
    """SSE 实时推送：每当有人增删改，所有在线协作者都会收到事件并自动刷新对应页面。"""
    async def gen():
        last = 0
        last_presence = 0.0
        while True:
            now = time.time()
            with _EVENT_LOCK:
                items = [e for e in _EVENTS if e["id"] > last]
            for e in items:
                last = e["id"]
                yield "id: %d\nretry: 3000\ndata: %s\n\n" % (
                    e["id"], json.dumps(e, ensure_ascii=False))
            # 每 5 秒推送一次在线成员快照（命名事件，前端单独监听）
            if now - last_presence >= 5:
                last_presence = now
                snap = {"type": "presence", "online": online_members()}
                yield "event: presence\ndata: %s\n\n" % json.dumps(snap, ensure_ascii=False)
            await asyncio.sleep(0.5)

    return StreamingResponse(
        gen(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "Connection": "keep-alive",
                 "X-Accel-Buffering": "no"},
    )


@app.post("/api/heartbeat")
async def heartbeat(request: Request):
    """客户端定时上报心跳，用于在线状态统计。"""
    uid = request.headers.get("X-User-Id")
    if uid:
        try:
            with _ONLINE_LOCK:
                _ONLINE[int(uid)] = time.time()
        except Exception:
            pass
    return {"ok": True}


@app.get("/api/presence")
def presence():
    return {"online": online_members()}


@app.get("/api/activity")
def activity(limit: int = 60):
    return db.recent_activity(limit)


@app.get("/")
def index():
    return FileResponse(INDEX)


# ---------------- 店铺 ----------------
@app.get("/api/shops")
def list_shops():
    c = db.conn()
    rows = c.execute("SELECT * FROM shops ORDER BY id").fetchall()
    c.close()
    return rows_to_list(rows)


# ---------------- 成员（进入系统身份登记） ----------------
@app.get("/api/members")
def list_members():
    return db.list_members()


@app.post("/api/members")
async def create_member(request: Request):
    b = await request.json()
    name = (b.get("name") or "").strip()
    department = (b.get("department") or "").strip()
    if not name or not department:
        raise HTTPException(400, detail="真实姓名与所在部门均为必填")
    if len(name) > 20 or len(department) > 30:
        raise HTTPException(400, detail="姓名/部门长度超限")
    m = db.add_member(name, department)
    broadcast("members", "join", f"{name} 加入了团队", None, m["id"],
              actor_id=m["id"], actor_name=name)
    return m


@app.delete("/api/members/{mid}")
def delete_member(mid: int, request: Request = None):
    r = db.delete_member(mid)
    broadcast("members", "leave", f"移除了成员 #{mid}", request, mid)
    return r



@app.post("/api/shops")
async def create_shop(request: Request):
    b = await request.json()
    c = db.conn()
    cur = c.execute(
        "INSERT INTO shops(name,owner,email,url,grp,role,price) VALUES(?,?,?,?,?,?,?)",
        (b.get("name"), b.get("owner"), b.get("email"), b.get("url"), b.get("grp"), b.get("role", "agent"), _f(b.get("price"))),
    )
    c.commit()
    rid = cur.lastrowid
    row = c.execute("SELECT * FROM shops WHERE id=?", (rid,)).fetchone()
    c.close()
    broadcast("shops", "create", f"新增店铺 {b.get('name')}", request, rid)
    return dict(row)


@app.put("/api/shops/{sid}")
async def update_shop(sid: int, request: Request):
    b = await request.json()
    c = db.conn()
    c.execute(
        "UPDATE shops SET name=?,owner=?,email=?,url=?,grp=?,role=?,price=? WHERE id=?",
        (b.get("name"), b.get("owner"), b.get("email"), b.get("url"), b.get("grp"), b.get("role", "agent"), _f(b.get("price")), sid),
    )
    c.commit()
    row = c.execute("SELECT * FROM shops WHERE id=?", (sid,)).fetchone()
    c.close()
    if not row:
        raise HTTPException(404)
    broadcast("shops", "update", f"编辑店铺 {b.get('name')}", request, sid)
    return dict(row)


@app.delete("/api/shops/{sid}")
def delete_shop(sid: int, request: Request = None):
    c = db.conn()
    c.execute("DELETE FROM shops WHERE id=?", (sid,))
    c.commit()
    c.close()
    broadcast("shops", "delete", f"删除店铺 #{sid}", request, sid)
    return {"ok": True}


@app.post("/api/shops/{sid}/image")
async def shop_image(sid: int, file: UploadFile = File(...), request: Request = None):
    name = save_upload(file)
    c = db.conn()
    c.execute("UPDATE shops SET image=? WHERE id=?", (name, sid))
    c.commit()
    c.close()
    broadcast("shops", "update", f"更新了店铺 #{sid} 主图", request, sid)
    return {"image": name}


def find_or_create_shop(c, url, role, name=None, price=None, image=None):
    row = c.execute("SELECT * FROM shops WHERE url=?", (url,)).fetchone()
    if row:
        if image and not row["image"]:
            c.execute("UPDATE shops SET image=? WHERE id=?", (image, row["id"]))
            c.commit()
        return dict(c.execute("SELECT * FROM shops WHERE id=?", (row["id"],)).fetchone())
    cur = c.execute(
        "INSERT INTO shops(name,owner,email,url,grp,role,price,image) VALUES(?,?,?,?,?,?,?,?)",
        (name or url[:40], "", "", url, "自动巡检", role, price, image),
    )
    c.commit()
    return dict(c.execute("SELECT * FROM shops WHERE id=?", (cur.lastrowid,)).fetchone())


# ---------------- SKU 映射 ----------------
@app.get("/api/sku")
def list_sku():
    c = db.conn()
    rows = c.execute("SELECT * FROM sku ORDER BY id").fetchall()
    c.close()
    return rows_to_list(rows)


@app.post("/api/sku")
async def create_sku(request: Request):
    b = await request.json()
    c = db.conn()
    cur = c.execute(
        "INSERT INTO sku(official_sku,official_url,agent_shop,agent_sku,agent_url,official_price,agent_price) VALUES(?,?,?,?,?,?,?)",
        (b.get("official_sku"), b.get("official_url"), b.get("agent_shop"), b.get("agent_sku"), b.get("agent_url"), _f(b.get("official_price")), _f(b.get("agent_price"))),
    )
    c.commit()
    rid = cur.lastrowid
    row = c.execute("SELECT * FROM sku WHERE id=?", (rid,)).fetchone()
    c.close()
    broadcast("sku", "create", f"新增 SKU 映射 {b.get('official_sku')}", request, rid)
    return dict(row)


@app.put("/api/sku/{kid}")
async def update_sku(kid: int, request: Request):
    b = await request.json()
    c = db.conn()
    c.execute(
        "UPDATE sku SET official_sku=?,official_url=?,agent_shop=?,agent_sku=?,agent_url=?,official_price=?,agent_price=? WHERE id=?",
        (b.get("official_sku"), b.get("official_url"), b.get("agent_shop"), b.get("agent_sku"), b.get("agent_url"), _f(b.get("official_price")), _f(b.get("agent_price")), kid),
    )
    c.commit()
    row = c.execute("SELECT * FROM sku WHERE id=?", (kid,)).fetchone()
    c.close()
    if not row:
        raise HTTPException(404)
    broadcast("sku", "update", f"编辑 SKU 映射 #{kid}", request, kid)
    return dict(row)


@app.delete("/api/sku/{kid}")
def delete_sku(kid: int, request: Request = None):
    c = db.conn()
    c.execute("DELETE FROM sku WHERE id=?", (kid,))
    c.commit()
    c.close()
    broadcast("sku", "delete", f"删除 SKU 映射 #{kid}", request, kid)
    return {"ok": True}


# ---------------- 视觉标准库 ----------------
@app.get("/api/standards")
def list_standards():
    c = db.conn()
    rows = c.execute("SELECT * FROM standards ORDER BY id").fetchall()
    c.close()
    return rows_to_list(rows)


@app.post("/api/standards")
async def create_standard(request: Request):
    b = await request.json()
    c = db.conn()
    cur = c.execute(
        "INSERT INTO standards(name,page_type,hex_color,logo_x,logo_y,logo_w,logo_h,forbidden_fonts) VALUES(?,?,?,?,?,?,?,?)",
        (b.get("name"), b.get("page_type"), b.get("hex_color"), _f(b.get("logo_x")), _f(b.get("logo_y")), _f(b.get("logo_w")), _f(b.get("logo_h")), b.get("forbidden_fonts")),
    )
    c.commit()
    rid = cur.lastrowid
    row = c.execute("SELECT * FROM standards WHERE id=?", (rid,)).fetchone()
    c.close()
    broadcast("standards", "create", f"新增视觉标准 {b.get('name')}", request, rid)
    return dict(row)


@app.put("/api/standards/{sid}")
async def update_standard(sid: int, request: Request):
    b = await request.json()
    c = db.conn()
    c.execute(
        "UPDATE standards SET name=?,page_type=?,hex_color=?,logo_x=?,logo_y=?,logo_w=?,logo_h=?,forbidden_fonts=? WHERE id=?",
        (b.get("name"), b.get("page_type"), b.get("hex_color"), _f(b.get("logo_x")), _f(b.get("logo_y")), _f(b.get("logo_w")), _f(b.get("logo_h")), b.get("forbidden_fonts"), sid),
    )
    c.commit()
    row = c.execute("SELECT * FROM standards WHERE id=?", (sid,)).fetchone()
    c.close()
    if not row:
        raise HTTPException(404)
    broadcast("standards", "update", f"编辑视觉标准 {b.get('name')}", request, sid)
    return dict(row)


@app.delete("/api/standards/{sid}")
def delete_standard(sid: int, request: Request = None):
    c = db.conn()
    c.execute("DELETE FROM standards WHERE id=?", (sid,))
    c.commit()
    c.close()
    broadcast("standards", "delete", f"删除视觉标准 #{sid}", request, sid)
    return {"ok": True}


@app.post("/api/standards/{sid}/image")
async def standard_image(sid: int, file: UploadFile = File(...), request: Request = None):
    name = save_upload(file)
    c = db.conn()
    c.execute("UPDATE standards SET image=? WHERE id=?", (name, sid))
    c.commit()
    c.close()
    broadcast("standards", "update", f"更新了视觉标准 #{sid} 基准图", request, sid)
    return {"image": name}


# ---------------- 设置 / 邮件模板 ----------------
@app.get("/api/settings")
def get_settings():
    c = db.conn()
    s = c.execute("SELECT k,v FROM settings").fetchall()
    t = c.execute("SELECT subject,body FROM email_template WHERE id=1").fetchone()
    c.close()
    out = {r["k"]: r["v"] for r in s}
    out["email_subject"] = t["subject"]
    out["email_body"] = t["body"]
    return out


@app.put("/api/settings")
async def put_settings(request: Request):
    b = await request.json()
    c = db.conn()
    for k in ("visual_threshold", "price_tolerance"):
        if k in b:
            c.execute("INSERT INTO settings(k,v) VALUES(?,?) ON CONFLICT(k) DO UPDATE SET v=?", (k, str(b[k]), str(b[k])))
    if "email_subject" in b or "email_body" in b:
        t = c.execute("SELECT subject,body FROM email_template WHERE id=1").fetchone()
        subj = b.get("email_subject", t["subject"])
        body = b.get("email_body", t["body"])
        c.execute("UPDATE email_template SET subject=?,body=? WHERE id=1", (subj, body))
    c.commit()
    c.close()
    broadcast("settings", "update", "更新了项目设置（阈值/通知模板）", request)
    return {"ok": True}



# ---------------- 精准对照表巡检（单链接 + 批量导入） ----------------
# 设计：放弃全店爬取，改为「精准对照表」——每条巡检只比对用户明确给出的
#   (代理商链接, 官旗标准链接, 图片类型) 这一对链接，杜绝商品匹配混乱。
# 抓取策略：对单条链接按「主图 / SKU缩略图列表 / 详情页长图」三类分库存储，
# 比对时同样分三类组对比，官网旗(左) vs 代理(右)，红框标注缺失/差异。


def _read_bytes(folder, names):
    """把分库里的一组文件名读成字节列表，交给 compare 做比对。"""
    out = []
    for n in (names or []):
        p = os.path.join(UPLOAD, folder, n)
        if os.path.exists(p):
            try:
                with open(p, "rb") as f:
                    out.append(f.read())
            except Exception:
                pass
    return out


def _insp_dict(row):
    d = dict(row)
    for k in ("off_main", "off_sku", "off_detail", "ag_main", "ag_sku", "ag_detail",
              "cmp_main", "cmp_sku", "cmp_detail"):
        v = d.get(k)
        if isinstance(v, str):
            try:
                d[k] = json.loads(v)
            except Exception:
                pass
    return d


def _host_of(url):
    try:
        from urllib.parse import urlparse
        return (urlparse(url or "").netloc or str(url)).replace("www.", "")
    except Exception:
        return str(url)


# ---------------------------------------------------------------------------
# 有界爬虫线程池：严格限制同时运行的浏览器实例数。
# 曾经所有巡检（单条新增 / 重抓 / 批量导入每一行）都直接 threading.Thread().start()，
# 导入几十行 Excel 会瞬间并发拉起几十×2 个 Edge 浏览器，把本机 CPU/内存打满，
# 导致同机运行的浏览器标签页直接"卡退"。改为统一入队、由固定数量的 worker 顺序执行。
# ---------------------------------------------------------------------------
CRAWLER_MAX_WORKERS = 2           # 同时最多 2 条巡检在跑（每条开 2 个浏览器，故上限 4 个）
CRAWLER_MAX_IMPORT_ROWS = 1000    # 单次导入行数上限，防止超大文件拖垮队列与数据库
CRAWLER_TIMEOUT = 240             # 单条抓取硬超时（秒）：到点强制置 error，避免永久挂死占用 worker
_crawler_queue = queue.Queue()
_crawler_started = False
_crawler_lock = threading.Lock()
CRAWLER_CANCEL = {}               # iid -> True：看门狗超时后置位，子线程完成后不再覆盖 error 状态


# ---------------------------------------------------------------------------
# 本地登录唤醒：用 Playwright 启动「有头」浏览器，引导用户手机淘宝扫码登录；
# 登录成功后把 storage_state 保存为 auth.json，后续抓取直接加载，无需重复手动复制 Cookie。
# Cookie 过期只需人工重新扫码一次，系统会自动更新本地缓存。
# ---------------------------------------------------------------------------
LOGIN_TIMEOUT = 180             # 扫码登录等待上限（秒）；二维码过期则提示重试
_login_lock = threading.Lock()
_login_state = {"running": False, "logged_in": False, "saved_at": None,
                "error": None, "message": "",
                "cookie_hits": [], "page_host": ""}  # 诊断字段：登录态 Cookie 命中名 / 当前页 host
_login_cancel = False      # 用户主动取消 / 重开时置位，登录线程检测到即退出
_login_browser = None       # 当前登录浏览器实例（用于取消时关闭窗口）
_login_thread = None        # 当前登录线程（用于重开前 join 等待退出）


def _abort_login():
    """中止进行中的登录流程：关闭浏览器窗口并置取消标志。"""
    global _login_browser, _login_cancel
    _login_cancel = True
    b = _login_browser
    if b is not None:
        try:
            b.close()
        except Exception:
            pass


@app.get("/api/taobao-login/status")
def taobao_login_status():
    st = dict(_login_state)
    # 只有「文件存在且经复核确含登录凭据」才算已登录，
    # 杜绝仅凭文件存在就显示“已登录（本地缓存）”的假象（旧的匿名态会被判为未登录）
    st["auth_file_exists"] = bool(os.path.exists(crawler.AUTH_PATH)
                                  and crawler._auth_has_login(crawler.AUTH_PATH))
    if st["auth_file_exists"] and not st["saved_at"]:
        try:
            st["saved_at"] = datetime.datetime.fromtimestamp(
                os.path.getmtime(crawler.AUTH_PATH)).isoformat()
        except Exception:
            pass
    # 诊断：若已登录/有缓存，补充登录态 Cookie 命中情况（仅名字，不含值），便于排查
    if st.get("auth_file_exists"):
        try:
            with open(crawler.AUTH_PATH, "r", encoding="utf-8") as f:
                _d = json.load(f)
            _n = {str(c.get("name", "")).lower() for c in (_d.get("cookies") or [])}
            st["cookie_hits"] = [k for k in ("tracknick", "x5sec", "sgcookie") if k in _n]
        except Exception:
            pass
    return st


@app.post("/api/taobao-login/start")
def taobao_login_start():
    global _login_thread
    # 若已有登录流程（含用户误关窗口后未退出的卡死进程），先中止并等待其退出，再重开
    if _login_state["running"]:
        _abort_login()
        if _login_thread is not None:
            _login_thread.join(timeout=5)
    with _login_lock:
        _login_state["running"] = True
        _login_state["logged_in"] = False
        _login_state["error"] = None
        _login_state["saved_at"] = None
        _login_state["message"] = "正在打开浏览器，请用手机淘宝扫码…"
    _login_thread = threading.Thread(target=_run_taobao_login, daemon=True)
    _login_thread.start()
    return {"ok": True}


@app.post("/api/taobao-login/cancel")
def taobao_login_cancel():
    _abort_login()
    return {"ok": True}


@app.post("/api/taobao-login/clear")
def taobao_login_clear():
    try:
        if os.path.exists(crawler.AUTH_PATH):
            os.remove(crawler.AUTH_PATH)
    except Exception:
        pass
    _login_state["logged_in"] = False
    _login_state["saved_at"] = None
    return {"ok": True}


def _run_taobao_login():
    global _login_browser, _login_cancel
    _login_cancel = False
    browser = None
    try:
        # 每次扫码都从干净状态开始：先删除旧的 auth.json，避免上一次（可能误判的）匿名态被复用
        if os.path.exists(crawler.AUTH_PATH):
            try:
                os.remove(crawler.AUTH_PATH)
            except Exception:
                pass
        from playwright.sync_api import sync_playwright
        with sync_playwright() as p:
            browser = crawler._launch_browser(p, headless=False)
            _login_browser = browser
            ctx = browser.new_context(
                viewport={"width": 1280, "height": 900},
                user_agent=crawler.UA, locale="zh-CN")
            page = ctx.new_page()
            _login_state["message"] = "正在打开淘宝首页，请用手机淘宝扫码…"
            try:
                page.goto("https://www.taobao.com",
                          wait_until="domcontentloaded", timeout=30000)
            except Exception:
                pass
            deadline = time.time() + LOGIN_TIMEOUT
            while time.time() < deadline:
                if _login_cancel:
                    _login_state["message"] = "已取消登录"
                    return
                # 用户手动关闭了浏览器窗口：立即结束，避免线程挂到超时
                alive = True
                try:
                    alive = browser.is_connected() and not page.is_closed()
                except Exception:
                    alive = False
                if not alive:
                    _login_state["error"] = "登录窗口已关闭，请重新点击扫码登录"
                    _login_state["message"] = "登录窗口已关闭"
                    return
                # 诊断信息：记录当前页 host 与命中的登录态 Cookie 名（仅名字，不含值）
                try:
                    _login_state["page_host"] = (urlparse(page.url).netloc or "")
                except Exception:
                    pass
                try:
                    _cnow = page.context.cookies() or []
                    _nnow = {str(c.get("name", "")).lower() for c in _cnow}
                    _login_state["cookie_hits"] = [
                        k for k in ("tracknick", "x5sec", "sgcookie") if k in _nnow]
                except Exception:
                    _login_state["cookie_hits"] = []
                try:
                    if crawler._is_taobao_logged_in(page):
                        break
                except Exception:
                    # 浏览器被关闭等异常，直接结束本次登录
                    _login_state["error"] = "登录窗口已关闭，请重新点击扫码登录"
                    _login_state["message"] = "登录窗口已关闭"
                    return
                time.sleep(2)
            else:
                # while 正常走完（超时未登录）
                _login_state["error"] = "登录超时（二维码过期或未扫码），请重试"
                _login_state["message"] = "登录超时，请重试"
                return
            if _login_cancel:
                return
            if crawler._is_taobao_logged_in(page):
                try:
                    ctx.storage_state(path=crawler.AUTH_PATH)
                except Exception as e:
                    _login_state["error"] = "保存登录状态失败：" + str(e)[:120]
                else:
                    # 关键复核：确认落盘的登录态真的含登录凭据，杜绝「假登录」
                    if crawler._auth_has_login(crawler.AUTH_PATH):
                        _login_state["logged_in"] = True
                        _login_state["saved_at"] = datetime.datetime.now().isoformat()
                        _login_state["message"] = "登录成功，已保存登录状态。"
                    else:
                        # 误判：页面未真正登录，删掉这份匿名态，要求重新扫码
                        try:
                            os.remove(crawler.AUTH_PATH)
                        except Exception:
                            pass
                        _login_state["logged_in"] = False
                        _login_state["error"] = "未检测到真实登录态（可能未真正扫码），请重新扫码"
                        _login_state["message"] = "未检测到登录态，请重新扫码"
            else:
                _login_state["error"] = "登录超时（二维码过期或未扫码），请重试"
                _login_state["message"] = "登录超时，请重试"
    except Exception as e:
        _login_state["error"] = str(e)[:200]
        _login_state["message"] = "登录流程出错：" + str(e)[:100]
    finally:
        _login_cancel = False
        _login_browser = None
        if browser:
            try:
                browser.close()
            except Exception:
                pass
        _login_state["running"] = False


def _crawler_worker():
    """常驻 worker：从队列取任务执行，单条异常不影响后续任务与线程存活。"""
    while True:
        item = _crawler_queue.get()
        try:
            if item is None:
                break
            iid, request = item
            try:
                _run_inspection(iid, request)
            except Exception as _e:  # noqa: BLE001
                print(f"[crawler] 巡检 #{iid} 执行异常：{_e}")
        finally:
            _crawler_queue.task_done()


def _ensure_crawler_pool():
    global _crawler_started
    with _crawler_lock:
        if _crawler_started:
            return
        for _ in range(CRAWLER_MAX_WORKERS):
            threading.Thread(target=_crawler_worker, daemon=True).start()
        _crawler_started = True


def enqueue_inspection(iid, request=None):
    """把巡检任务放入有界线程池排队执行（替代直接 threading.Thread().start()）。"""
    _ensure_crawler_pool()
    _crawler_queue.put((iid, request))


def _run_inspection_core(iid, request=None):
    """后台线程（由 _run_inspection 在子线程中调用）：对单条记录真实抓取并比对。"""
    actor_id, actor_name = _actor(request)
    c = db.conn()
    c.execute("UPDATE inspections SET state='running' WHERE id=?", (iid,))
    c.commit()
    row = c.execute("SELECT * FROM inspections WHERE id=?", (iid,)).fetchone()
    c.close()
    if not row:
        return
    agent_url = row["agent_url"]
    official_url = row["official_url"]
    only = (row["image_type"] or "").strip()  # "" 表示全部分类
    cookies = row["cookies"]
    try:
        from playwright.sync_api import sync_playwright
        with sync_playwright() as _p:
            browser = crawler._launch_browser(_p)
            # 复用「本地登录唤醒」保存的淘宝/天猫登录态（auth.json），免手动填 Cookie
            # 仅当文件存在且经复核确含登录凭据时使用，避免误用匿名态导致抓取被登录墙拦截
            auth = (crawler.AUTH_PATH
                    if (not cookies and os.path.exists(crawler.AUTH_PATH)
                        and crawler._auth_has_login(crawler.AUTH_PATH))
                    else None)
            try:
                off = crawler.crawl_product(official_url, UPLOAD, cookies=cookies, browser=browser, storage_state=auth)
                ag = crawler.crawl_product(agent_url, UPLOAD, cookies=cookies, browser=browser, storage_state=auth)
            finally:
                try:
                    browser.close()
                except Exception:
                    pass
        thr = float(db.get_setting("visual_threshold", 90))
        off_main = _read_bytes("main", off["main"])
        ag_main = _read_bytes("main", ag["main"])
        off_sku = _read_bytes("sku", off["sku"])
        ag_sku = _read_bytes("sku", ag["sku"])
        off_det = _read_bytes("detail", off["detail"])
        ag_det = _read_bytes("detail", ag["detail"])

        cmp_main = {"skipped": True} if (only and only != "main") else \
            compare.compare_main(off_main, ag_main, thr)
        cmp_sku = {"skipped": True} if (only and only != "sku") else \
            compare.compare_sku(off_sku, ag_sku, thr)
        cmp_detail = {"skipped": True} if (only and only != "detail") else \
            compare.compare_detail(off_det, ag_det)

        violation = (
            (isinstance(cmp_main, dict) and cmp_main.get("has_diff") and not cmp_main.get("skipped")) or
            (isinstance(cmp_sku, dict) and cmp_sku.get("mismatch") and not cmp_sku.get("skipped")) or
            (isinstance(cmp_detail, dict) and cmp_detail.get("mismatch") and not cmp_detail.get("skipped"))
        )
        status = "pending" if violation else "modified"
        simulated = off.get("simulated") or ag.get("simulated")
        login_wall = off.get("login_wall") or ag.get("login_wall")
        if login_wall:
            need = []
            if off.get("login_wall"):
                need.append("官旗标准链接")
            if ag.get("login_wall"):
                need.append("代理商链接")
            note = ("以下链接被登录墙拦截（淘宝/天猫反爬，需登录态才能查看商品图），"
                    "请在表单填写对应 Cookie 后点击「重新抓取复核」：" + "、".join(need))
        elif simulated:
            bad = []
            if off.get("simulated"):
                bad.append("官旗标准链接")
            if ag.get("simulated"):
                bad.append("代理商链接")
            note = "以下链接未能真实抓取、已用演示图兜底：" + "、".join(bad) + "（请检查链接可达性 / 本机外网）"
        else:
            note = None
        # 若看门狗已超时将此巡检置为 error，则不覆盖其最终状态
        if CRAWLER_CANCEL.get(iid):
            return
        c = db.conn()
        c.execute(
            """UPDATE inspections SET state='done', status=?, shop_host=?,
               off_main=?, off_sku=?, off_detail=?, ag_main=?, ag_sku=?, ag_detail=?,
               cmp_main=?, cmp_sku=?, cmp_detail=?, price_official=?, price_agent=?,
               error=?, updated_at=CURRENT_TIMESTAMP WHERE id=?""",
            (status, _host_of(agent_url),
             json.dumps(off["main"]), json.dumps(off["sku"]), json.dumps(off["detail"]),
             json.dumps(ag["main"]), json.dumps(ag["sku"]), json.dumps(ag["detail"]),
             json.dumps(cmp_main), json.dumps(cmp_sku), json.dumps(cmp_detail),
             off.get("price"), ag.get("price"), note, iid))
        c.commit(); c.close()
        broadcast("inspections", "update", f"完成单链接巡检 #{iid}", request, iid,
                  actor_id=actor_id, actor_name=actor_name)
    except Exception as e:
        c = db.conn()
        c.execute("UPDATE inspections SET state='error', error=? WHERE id=?", (str(e)[:300], iid))
        c.commit(); c.close()
        broadcast("inspections", "update", f"单链接巡检 #{iid} 失败：{str(e)[:80]}", request, iid,
                  actor_id=actor_id, actor_name=actor_name)


def _run_inspection(iid, request=None):
    """有界线程池 worker 调用的入口：在子线程中执行抓取，主 worker 用 join(timeout) 兜底。

    任一条巡检（真实抓取）若因链接不可达、被反爬拦截、网络过慢等原因长时间挂死，
    硬超时（CRAWLER_TIMEOUT）后会强制将其置为 error，释放卡片与 worker，
    避免整条队列永远停在「抓取中」把有界线程池拖死。
    """
    actor_id, actor_name = _actor(request)
    c = db.conn()
    c.execute("UPDATE inspections SET state='running' WHERE id=?", (iid,))
    c.commit(); c.close()
    CRAWLER_CANCEL.pop(iid, None)

    holder = {}
    def _crawl():
        try:
            _run_inspection_core(iid, request)
        except Exception as _e:  # noqa: BLE001
            holder["err"] = _e
    th = threading.Thread(target=_crawl, daemon=True)
    th.start()
    th.join(CRAWLER_TIMEOUT)
    if th.is_alive():
        # 抓取超时仍未结束：强制置 error，并通知看门狗让子线程完成后不要覆盖状态
        CRAWLER_CANCEL[iid] = True
        c = db.conn()
        c.execute("UPDATE inspections SET state='error', error=? WHERE id=? AND state='running'",
                  (f"抓取超时（{CRAWLER_TIMEOUT} 秒未结束）。该链接可能不可达、被反爬拦截或网络过慢；"
                   "建议检查链接，或在单链接巡检表单填写「淘宝/天猫登录态 Cookie」后点击「重新抓取复核」", iid))
        c.commit(); c.close()
        broadcast("inspections", "update", f"巡检 #{iid} 抓取超时，已标记为失败", request, iid,
                  actor_id=actor_id, actor_name=actor_name)
        return
    if "err" in holder:
        c = db.conn()
        c.execute("UPDATE inspections SET state='error', error=? WHERE id=? AND state='running'",
                  (str(holder["err"])[:300], iid))
        c.commit(); c.close()


@app.post("/api/inspections")
async def create_inspection(request: Request):
    b = await request.json()
    agent_url = (b.get("agent_url") or "").strip()
    official_url = (b.get("official_url") or "").strip()
    image_type = (b.get("image_type") or "").strip() or None
    title = (b.get("title") or "").strip() or None
    cookies = (b.get("cookies") or "").strip() or None
    if not agent_url or not official_url:
        raise HTTPException(400, detail="代理商链接与官旗标准链接均为必填")
    row = db.create_inspection(agent_url, official_url, image_type, title,
                               by_member=_actor(request)[0], cookies=cookies)
    scope = ("仅 " + image_type) if image_type else "主图 / SKU / 详情 全部分类"
    broadcast("inspections", "create", f"新增单链接巡检（{scope}）", request, row["id"])
    enqueue_inspection(row["id"], request)
    return {"id": row["id"]}


@app.get("/api/inspections")
def list_inspections(status: str = None, state: str = None):
    return [_insp_dict(r) for r in db.list_inspections(status, state)]


@app.get("/api/inspections/export")
def export_inspections():
    """导出全部巡检为 CSV（含核销留痕字段），供数据报表离线分析。"""
    import csv
    c = db.conn()
    rows = [dict(r) for r in c.execute("SELECT * FROM inspections ORDER BY id DESC").fetchall()]
    c.close()
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["ID", "标题", "代理商链接", "官旗链接", "图片类型", "处理状态", "抓取状态",
                "主图相似度", "SKU差异", "详情缺失数", "备注", "核销人", "核销时间", "创建时间"])
    for r in rows:
        cm = _safe_json(r.get("cmp_main"))
        cs = _safe_json(r.get("cmp_sku"))
        cd = _safe_json(r.get("cmp_detail"))
        miss = cd.get("missing_in_agent") if isinstance(cd, dict) else None
        w.writerow([
            r["id"], r.get("title") or "", r.get("agent_url"), r.get("official_url"),
            r.get("image_type") or "全部", r.get("status"), r.get("state"),
            cm.get("score", "") if isinstance(cm, dict) else "",
            ("是" if isinstance(cs, dict) and cs.get("mismatch") else "否"),
            (len(miss) if isinstance(miss, list) else 0),
            r.get("note") or "", r.get("verified_by_name") or "", r.get("verified_at") or "",
            r.get("created_at"),
        ])
    return Response(content=buf.getvalue().encode("utf-8-sig"),
                    media_type="text/csv; charset=utf-8",
                    headers={"Content-Disposition": "attachment; filename=inspections_export.csv"})


@app.get("/api/inspections/{iid}")
def get_inspection(iid: int):
    row = db.get_inspection(iid)
    if not row:
        raise HTTPException(404)
    return _insp_dict(row)


@app.patch("/api/inspections/{iid}")
async def patch_inspection(iid: int, request: Request):
    b = await request.json()
    fields = {}
    for k in ("status", "note", "annotation"):
        if k in b and b[k] is not None:
            fields[k] = b[k]
    row = db.update_inspection(iid, **fields)
    if not row:
        raise HTTPException(404)
    detail = "更新了巡检处理"
    if b.get("status"):
        detail = f"将巡检 #{iid} 状态改为 {b['status']}"
    broadcast("inspections", "update", detail, request, iid)
    return _insp_dict(row)


@app.delete("/api/inspections/{iid}")
def delete_inspection(iid: int, request: Request = None):
    db.delete_inspection(iid)
    broadcast("inspections", "delete", f"删除巡检记录 #{iid}", request, iid)
    return {"ok": True}


@app.post("/api/inspections/{iid}/recrawl")
async def recrawl_inspection(iid: int, request: Request = None):
    """重新抓取复核：对同一条对照表记录重新真实抓取两类链接的三类图片并比对。"""
    if not db.get_inspection(iid):
        raise HTTPException(404)
    broadcast("inspections", "update", f"重新抓取复核巡检 #{iid}", request, iid)
    enqueue_inspection(iid, request)
    return {"id": iid, "status": "submitted"}


def _save_to_folder(folder, file: UploadFile) -> str:
    os.makedirs(os.path.join(UPLOAD, folder), exist_ok=True)
    ext = os.path.splitext(file.filename or "img")[1].lower() or ".png"
    name = uuid.uuid4().hex + ext
    with open(os.path.join(UPLOAD, folder, name), "wb") as f:
        f.write(file.file.read())
    return name


@app.post("/api/inspections/{iid}/recheck")
async def recheck_inspection(iid: int, file: UploadFile = File(...), request: Request = None):
    """手动上传代理商主图做复核：用新图与官旗主图重新比对（无需重新抓取）。"""
    name = _save_to_folder("main", file)
    c = db.conn()
    row = c.execute("SELECT * FROM inspections WHERE id=?", (iid,)).fetchone()
    if not row:
        c.close(); raise HTTPException(404)
    thr = float(db.get_setting("visual_threshold", 90))
    off = _read_bytes("main", json.loads(row["off_main"] or "[]"))
    with open(os.path.join(UPLOAD, "main", name), "rb") as f:
        ag = f.read()
    cmp_main = compare.compare_pair(off[0] if off else None, ag, thr)
    status = "pending" if cmp_main.get("has_diff") else "modified"
    c.execute(
        "UPDATE inspections SET ag_main=?, cmp_main=?, status=?, updated_at=CURRENT_TIMESTAMP WHERE id=?",
        (json.dumps([name]), json.dumps(cmp_main), status, iid))
    c.commit(); c.close()
    broadcast("inspections", "update",
              f"上传复核图整改巡检 #{iid}（相似度 {cmp_main.get('score')}%）", request, iid)
    return _insp_dict(db.get_inspection(iid))


@app.post("/api/inspections/{iid}/verify")
async def verify_inspection(iid: int, file: UploadFile = File(...),
                            note: str = Form(""), request: Request = None):
    """核销（独立终态）：上传整改凭证截图 + 备注，标记「已核销」并留痕（谁/何时/凭证）。"""
    row = db.get_inspection(iid)
    if not row:
        raise HTTPException(404)
    name = _save_to_folder("evidence", file)
    actor_id, actor_name = _actor(request)
    c = db.conn()
    c.execute(
        "UPDATE inspections SET status='verified', evidence_path=?, verified_at=CURRENT_TIMESTAMP, "
        "verified_by=?, verified_by_name=?, note=?, updated_at=CURRENT_TIMESTAMP WHERE id=?",
        (name, actor_id, actor_name, (note or "").strip() or row.get("note"), iid))
    c.commit(); c.close()
    db.log_activity(actor_id, actor_name, "verify", "inspections",
                    f"核销巡检 #{iid}（凭证: {name}）", iid)
    broadcast("inspections", "verify", f"核销巡检 #{iid}", request, iid,
              actor_id=actor_id, actor_name=actor_name)
    return _insp_dict(db.get_inspection(iid))


@app.post("/api/inspections/{iid}/reopen")
async def reopen_inspection(iid: int, request: Request = None):
    """撤回核销：将「已核销」退回「已修改」，便于复核不通过时重新跟进。"""
    if not db.get_inspection(iid):
        raise HTTPException(404)
    row = db.update_inspection(iid, status="modified")
    actor_id, actor_name = _actor(request)
    db.log_activity(actor_id, actor_name, "reopen", "inspections", f"撤回核销巡检 #{iid}", iid)
    broadcast("inspections", "update", f"撤回核销巡检 #{iid}", request, iid,
              actor_id=actor_id, actor_name=actor_name)
    return _insp_dict(row)


@app.get("/api/template")
def download_template():
    """后台提供 Excel 精准对照表模板下载（字段：代理商链接 / 官旗标准链接 / 图片类型）。"""
    import openpyxl
    from openpyxl.utils import get_column_letter
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "精准对照表"
    ws.append(["代理商链接", "官旗标准链接", "图片类型（选填：main/sku/detail，留空=全部分类）"])
    ws.append(["https://agent.example.com/item/123", "https://flagship.example.com/item/123", ""])
    ws.append(["https://agent.example.com/item/456", "https://flagship.example.com/item/456", "sku"])
    for i, w in enumerate((42, 42, 46), 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    path = os.path.join(UPLOAD, "template.xlsx")
    wb.save(path)
    return FileResponse(path, filename="精准对照表模板.xlsx",
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.post("/api/inspections/import")
async def import_inspections(file: UploadFile = File(...), request: Request = None):
    """批量导入：解析 Excel 精准对照表，为每一行创建单链接巡检（复用同一抓取/比对管线）。"""
    import io, openpyxl
    data = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data))
    except Exception as e:
        raise HTTPException(400, detail="无法解析 Excel 文件：" + str(e))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"created": 0, "ids": []}
    data_rows = rows[1:]  # 跳过表头
    truncated = False
    if len(data_rows) > CRAWLER_MAX_IMPORT_ROWS:
        data_rows = data_rows[:CRAWLER_MAX_IMPORT_ROWS]
        truncated = True
    created = []
    for r in data_rows:
        if not r:
            continue
        agent = str(r[0] or "").strip() if len(r) > 0 else ""
        official = str(r[1] or "").strip() if len(r) > 1 else ""
        itype = str(r[2] or "").strip() if len(r) > 2 else ""
        if not agent or not official:
            continue
        row = db.create_inspection(agent, official, itype or None, None)
        created.append(row["id"])
    for iid in created:
        enqueue_inspection(iid, request)
    if created:
        broadcast("inspections", "create", f"批量导入 {len(created)} 条巡检", request)
    resp = {"created": len(created), "ids": created}
    if truncated:
        resp["note"] = (f"已超出单次导入上限 {CRAWLER_MAX_IMPORT_ROWS} 行，"
                        f"仅导入前 {CRAWLER_MAX_IMPORT_ROWS} 条；请分批上传。")
    return resp


@app.post("/api/batch-notify")
async def batch_notify(request: Request):
    b = await request.json()
    ids = b.get("ids", [])
    c = db.conn()
    t = c.execute("SELECT subject,body FROM email_template WHERE id=1").fetchone()
    out = []
    for iid in ids:
        insp = c.execute("SELECT * FROM inspections WHERE id=?", (iid,)).fetchone()
        if not insp:
            continue
        d = _insp_dict(insp)
        parts = []
        cm = d.get("cmp_main") or {}
        cs = d.get("cmp_sku") or {}
        cd = d.get("cmp_detail") or {}
        if cm.get("has_diff"):
            parts.append(f"主图相似度 {cm.get('score')}%")
        if cs.get("mismatch"):
            parts.append(f"SKU图缺失/差异 {len(cs.get('missing_in_agent', []))} 项")
        if cd.get("mismatch"):
            parts.append(f"详情图缺失 {len(cd.get('missing_in_agent', []))} 张")
        detail = "；".join(parts) or "整体视觉一致"
        shop = d.get("title") or d.get("agent_url") or "该店铺"
        subject = t["subject"].replace("{shop}", shop)
        body = (t["body"].replace("{shop}", shop).replace("{sku}", shop)
                .replace("{type}", "视觉合规").replace("{detail}", detail))
        out.append({"shop": shop, "subject": subject, "body": body})
        c.execute("UPDATE inspections SET status='notified' WHERE id=?", (iid,))
    c.commit(); c.close()
    broadcast("inspections", "notify", f"批量发送了 {len(ids)} 条整改通知", request)
    return out


def _safe_json(v):
    if isinstance(v, str):
        try:
            return json.loads(v)
        except Exception:
            return {}
    return v or {}


@app.get("/api/stats")
def stats():
    """数据报表统计：总量、各状态/各抓取态分布、合格率、近 14 天趋势、Top 店铺。"""
    c = db.conn()
    total = c.execute("SELECT COUNT(*) n FROM inspections").fetchone()["n"]
    by_status, by_state = {}, {}
    for s in ("pending", "notified", "modified", "verified"):
        by_status[s] = c.execute("SELECT COUNT(*) n FROM inspections WHERE status=?", (s,)).fetchone()["n"]
    for s in ("queued", "running", "done", "error"):
        by_state[s] = c.execute("SELECT COUNT(*) n FROM inspections WHERE state=?", (s,)).fetchone()["n"]
    diff_count = c.execute(
        "SELECT COUNT(*) n FROM inspections WHERE state='done' AND status='pending'").fetchone()["n"]
    consistent = c.execute(
        "SELECT COUNT(*) n FROM inspections WHERE state='done' AND status!='pending'").fetchone()["n"]
    trend = [dict(r) for r in c.execute(
        "SELECT substr(created_at,1,10) d, COUNT(*) n FROM inspections "
        "WHERE created_at >= date('now','-13 days') GROUP BY d ORDER BY d").fetchall()]
    top_shops = [dict(r) for r in c.execute(
        "SELECT shop_host, COUNT(*) n FROM inspections WHERE shop_host<>'' "
        "GROUP BY shop_host ORDER BY n DESC LIMIT 10").fetchall()]
    c.close()
    resolved = by_status.get("modified", 0) + by_status.get("verified", 0)
    pass_rate = round(resolved / total * 100, 1) if total else 0.0
    return {
        "total": total, "by_status": by_status, "by_state": by_state,
        "diff_count": diff_count, "consistent": consistent,
        "pass_rate": pass_rate, "trend": trend, "top_shops": top_shops,
    }


def _f(v):
    try:
        return float(v) if v not in (None, "") else None
    except (TypeError, ValueError):
        return None


if __name__ == "__main__":
    import os
    host = os.environ.get("HOST", "0.0.0.0")
    port = int(os.environ.get("PORT", "8000"))
    uvicorn.run(app, host=host, port=port, reload=False)
