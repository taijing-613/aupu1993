# -*- coding: utf-8 -*-
"""生成《淘宝视觉合规巡检系统》当前进度 Excel。"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ---- 通用样式 ----
HEAD_FILL = PatternFill("solid", fgColor="1F4E78")
HEAD_FONT = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name="微软雅黑", bold=True, size=14, color="1F4E78")
SUB_FONT = Font(name="微软雅黑", size=9, color="595959", italic=True)
BODY_FONT = Font(name="微软雅黑", size=10)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

STATUS_FILL = {
    "已完成": PatternFill("solid", fgColor="C6EFCE"),
    "进行中": PatternFill("solid", fgColor="FFEB9C"),
    "阻塞":   PatternFill("solid", fgColor="FFC7CE"),
    "未开始": PatternFill("solid", fgColor="F2F2F2"),
}
STATUS_FONT = {
    "已完成": Font(name="微软雅黑", size=10, color="006100"),
    "进行中": Font(name="微软雅黑", size=10, color="9C6500"),
    "阻塞":   Font(name="微软雅黑", size=10, color="9C0006"),
    "未开始": Font(name="微软雅黑", size=10, color="808080"),
}

def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = CENTER
        cell.border = BORDER

def grid(ws, r1, r2, c1, c2):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).border = BORDER
            if ws.cell(row=r, column=c).font is None or ws.cell(row=r, column=c).font.name is None:
                ws.cell(row=r, column=c).font = BODY_FONT

# =====================================================================
# Sheet 1: 模块进度总览
# =====================================================================
ws = wb.active
ws.title = "模块进度总览"
ws["A1"] = "淘宝视觉合规巡检系统 — 模块进度总览"
ws["A1"].font = TITLE_FONT
ws["A2"] = ("统计口径：以代码实际完成情况为准（含已补全的「核销留痕」「数据报表」）。"
            "数据截至 2026-08-21，Git 基线 master（本地 2 commits，待 push 远程）。")
ws["A2"].font = SUB_FONT

headers = ["序号", "功能模块", "完成度", "状态", "状态说明"]
hr = 4
for i, h in enumerate(headers, 1):
    ws.cell(row=hr, column=i, value=h)
style_header(ws, hr, len(headers))

modules = [
    (1, "店铺管理（增删改查）", "100%", "已完成", "完整 CRUD + 搜索/筛选，迁移自旧库已回填 shop_id。"),
    (2, "视觉标准库（标准图录入/编辑）", "100%", "已完成", "标准图上传 / 编辑 / 关联 SKU，支持官旗图高清化。"),
    (3, "巡检任务（单链接巡检）", "100%", "已完成", "单链接提交 → Playwright 抓取 → AI 比对 → 出结果。"),
    (4, "巡检任务（Excel 批量导入）", "100%", "已完成", "openpyxl 解析 + 模板导出，列校验与错误提示。"),
    (5, "图片自动抓取（Playwright + Cookie 注入）", "85%", "进行中", "代码完整、扫码登录真实有效；真实淘宝端到端召回率待联网回归验证。"),
    (6, "AI 视觉比对（SSIM / pHash 算法）", "100%", "已完成", "主图轮播 / SKU 色卡 / 详情 iframe 三类分抓比对，相似度 + 差异标注。"),
    (7, "违规看板（问题列表展示）", "100%", "已完成", "问题列表、按状态筛选、一键跳转对比视图。"),
    (8, "左右对比视图（官旗 vs 代理）", "100%", "已完成", "三栏对比 + 差异高亮，支持重抓 / 上传复核图。"),
    (9, "状态流转（待处理→已通知→已修改→已核销）", "100%", "已完成", "已恢复四态，含独立「已核销」态（早期曾并入已修改，本轮修复）。"),
    (10, "核销操作（上传截图凭证）", "100%", "已完成", "独立核销态 + 凭证截图落盘（uploads/evidence）+ 操作审计留痕（activity_log）。"),
    (11, "数据报表", "100%", "已完成", "统计接口（按状态/店铺/日期）+ SVG 图表 + CSV 导出（含核销留痕字段）。"),
    (12, "公网部署配置", "90%", "进行中", "Docker / Railway / Render 配置已就绪、gunicorn 生产启动已启用；待推送平台取得公网 URL。"),
    (13, "代码版本控制（Git）", "100%", "已完成", "本地仓库已建（master，2 commits，.gitignore 已排敏）；待 push 内网远程。"),
]

r = hr + 1
for m in modules:
    for c, v in enumerate(m, 1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.alignment = LEFT if c in (2, 5) else CENTER
        if c == 3:
            # 完成度上色
            pct = int(v.strip("%"))
            fill = ("C6EFCE" if pct >= 100 else "FFEB9C" if pct >= 80 else "FFC7CE")
            cell.fill = PatternFill("solid", fgColor=fill)
            cell.font = Font(name="微软雅黑", size=10, bold=True)
        if c == 4:
            cell.fill = STATUS_FILL.get(v, PatternFill())
            cell.font = STATUS_FONT.get(v, BODY_FONT)
    r += 1

grid(ws, hr, r - 1, 1, len(headers))

# 汇总行
sum_row = r
ws.cell(row=sum_row, column=2, value="整体加权完成度（不含 Git/部署）").font = Font(name="微软雅黑", size=10, bold=True)
# 计算前 11 个业务模块均值
valid = [int(m[2].strip("%")) for m in modules[:11]]
avg = round(sum(valid) / len(valid))
ws.cell(row=sum_row, column=3, value=f"{avg}%").font = Font(name="微软雅黑", size=10, bold=True)
ws.cell(row=sum_row, column=3).alignment = CENTER
ws.cell(row=sum_row, column=3).fill = PatternFill("solid", fgColor="C6EFCE")
ws.cell(row=sum_row, column=4, value="MVP 基本达成").font = Font(name="微软雅黑", size=10, bold=True)
ws.cell(row=sum_row, column=4).fill = STATUS_FILL["已完成"]
ws.cell(row=sum_row, column=4).alignment = CENTER

widths = [6, 38, 10, 10, 60]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A5"

# =====================================================================
# Sheet 2: 图片抓取专项
# =====================================================================
ws2 = wb.create_sheet("图片抓取专项")
ws2["A1"] = "图片抓取（核心难点）进展核对"
ws2["A1"].font = TITLE_FONT
ws2["A2"] = "对应交接需求中的「三、关于图片抓取」四项确认点。"
ws2["A2"].font = SUB_FONT

h2 = ["确认项", "状态", "说明"]
for i, h in enumerate(h2, 1):
    ws2.cell(row=4, column=i, value=h)
style_header(ws2, 4, len(h2))

rows2 = [
    ("Cookie 管理后台是否已开发完成？", "已完成",
     "双机制：① 扫码免 Cookie（手机淘宝扫码 → 保存 auth.json）；② 单条 Cookie 粘贴框（兼容无头环境）。"),
    ("Playwright 抓取脚本是否已调通？", "代码完成 / 待联网回归",
     "脚本逻辑闭环、合成测试通过、本机扫码登录态真实有效（cookie_hits: tracknick/sgcookie）；真实淘宝页面端到端召回率需联网回归。"),
    ("抓取失败降级为「待人工补图」逻辑？", "已实现",
     "登录墙检测 + 重抓 + 上传复核图三重兜底；无法自动抓取时转为人工补图。"),
    ("是否踩到新反爬坑（滑块/限流）？", "部分",
     "已解决「登录墙」；滑块验证 / IP 限流尚未自动处理，目前靠「过期 → 重扫一次」人工兜底。"),
]
r = 5
for row in rows2:
    for c, v in enumerate(row, 1):
        cell = ws2.cell(row=r, column=c, value=v)
        cell.alignment = LEFT if c != 2 else CENTER
        if c == 2:
            if "已完成" in v or "实现" in v:
                cell.fill = STATUS_FILL["已完成"]; cell.font = STATUS_FONT["已完成"]
            else:
                cell.fill = STATUS_FILL["进行中"]; cell.font = STATUS_FONT["进行中"]
    r += 1
grid(ws2, 4, r - 1, 1, len(h2))
ws2.column_dimensions["A"].width = 42
ws2.column_dimensions["B"].width = 20
ws2.column_dimensions["C"].width = 70
ws2.freeze_panes = "A5"

# =====================================================================
# Sheet 3: 交接材料清单
# =====================================================================
ws3 = wb.create_sheet("交接材料清单")
ws3["A1"] = "交接所需材料清单（提供状态）"
ws3["A1"].font = TITLE_FONT
ws3["A2"] = "供新同事快速接手，逐项核对。"
ws3["A2"].font = SUB_FONT

h3 = ["材料项", "是否就绪", "位置 / 说明"]
for i, h in enumerate(h3, 1):
    ws3.cell(row=4, column=i, value=h)
style_header(ws3, 4, len(h3))

rows3 = [
    ("项目目录结构说明", "已就绪", "见 HANDOVER.md 第六节 + app/ 源码（FastAPI+SQLite+SPA）。"),
    ("环境依赖清单", "已就绪", "app/requirements.txt（已补全 playwright/openpyxl/requests，并锁定验证环境版本）。"),
    ("本地运行步骤", "已就绪", "HANDOVER.md 第六节：建 venv → pip install → playwright install → pythonw app.py。"),
    ("配置文件说明（环境变量）", "已就绪", "HOST/PORT/DATA_DIR 经环境变量配置，默认值 0.0.0.0 / 8000 / app 目录。"),
    ("代码仓库", "本地就绪 / 待远程", "本地 master（2 commits）；push：git remote add origin <地址> && git push -u origin master。"),
    ("README / 部署文档", "已就绪", "app/DEPLOY.md（Docker / Railway / Render 三种方式 + 并发注意点）。"),
    ("数据库 ER / 表结构文档", "代码即文档", "app/db.py 含建表 SQL 注释；核心表：shops / standards / inspections / issues / members / activity_log。"),
    ("接口文档", "已就绪（自动）", "FastAPI Swagger：服务启动后访问 /docs（当前 http://127.0.0.1:8000/docs）。"),
    ("测试账号 / 测试数据", "无独立测试账号", "门禁为「填姓名+部门」即登记；可用真实淘宝链接做回归。"),
    ("公网访问地址", "待部署", "部署到 Railway/Render 后获取；当前仅本地 http://127.0.0.1:8000。"),
]
r = 5
for row in rows3:
    for c, v in enumerate(row, 1):
        cell = ws3.cell(row=r, column=c, value=v)
        cell.alignment = LEFT if c != 2 else CENTER
        if c == 2:
            if "已就绪" in v or "文档" in v:
                cell.fill = STATUS_FILL["已完成"]; cell.font = STATUS_FONT["已完成"]
            elif "待" in v or "无" in v:
                cell.fill = STATUS_FILL["进行中"]; cell.font = STATUS_FONT["进行中"]
            else:
                cell.fill = STATUS_FILL["已完成"]; cell.font = STATUS_FONT["已完成"]
    r += 1
grid(ws3, 4, r - 1, 1, len(h3))
ws3.column_dimensions["A"].width = 30
ws3.column_dimensions["B"].width = 18
ws3.column_dimensions["C"].width = 72
ws3.freeze_panes = "A5"

# =====================================================================
# Sheet 4: 风险与待办
# =====================================================================
ws4 = wb.create_sheet("风险与待办")
ws4["A1"] = "风险项与剩余待办"
ws4["A1"].font = TITLE_FONT
ws4["A2"] = "按优先级排列，供交接排期参考。"
ws4["A2"].font = SUB_FONT

h4 = ["优先级", "事项", "风险 / 说明", "建议动作"]
for i, h in enumerate(h4, 1):
    ws4.cell(row=4, column=i, value=h)
style_header(ws4, 4, len(h4))

rows4 = [
    ("P0", "推送至内网 Git 远程", "当前仅本地仓库，交接即「拷文件夹」，易丢版本。", "提供 Git 地址后执行 remote add + push。"),
    ("P0", "公网部署取得 URL", "「所有人可访问」尚未落地（CloudStudio 仅支持静态，本项目需 Python 托管）。", "用 Railway/Render（配置已就绪）部署，或自有服务器。"),
    ("P1", "图片抓取联网回归", "真实淘宝页面召回率为沙箱无法验证项。", "联网环境跑 3~5 真实链接，核对主图/SKU/详情抓取质量。"),
    ("P1", "滑块验证 / IP 限流", "反爬极端情形未自动处理，靠过期重扫兜底。", "观察频次；如高频触发再引入代理/打码服务。"),
    ("P2", "服务常驻稳定性", "本机曾因重复守护进程致只读库崩溃；当前为单实例稳定运行。", "部署到平台由平台保活；本机可选单 launcher（.guard.lock 防重复）。"),
    ("P2", "门禁体验优化", "开场强制填姓名+部门，报错已改显原文。", "如嫌繁琐可改为可选登记或关闭门禁。"),
]
PRIO_FILL = {"P0": "FFC7CE", "P1": "FFEB9C", "P2": "DDEBF7"}
r = 5
for row in rows4:
    for c, v in enumerate(row, 1):
        cell = ws4.cell(row=r, column=c, value=v)
        cell.alignment = LEFT if c in (2, 3, 4) else CENTER
        if c == 1:
            cell.fill = PatternFill("solid", fgColor=PRIO_FILL.get(v, "FFFFFF"))
            cell.font = Font(name="微软雅黑", size=10, bold=True)
    r += 1
grid(ws4, 4, r - 1, 1, len(h4))
ws4.column_dimensions["A"].width = 10
ws4.column_dimensions["B"].width = 24
ws4.column_dimensions["C"].width = 52
ws4.column_dimensions["D"].width = 46
ws4.freeze_panes = "A5"

out = r"C:\Users\J\WorkBuddy\2026-08-17-08-58-49\progress_20260821.xlsx"
wb.save(out)
print("SAVED:", out)
print("Sheets:", wb.sheetnames)
