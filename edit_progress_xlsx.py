# -*- coding: utf-8 -*-
"""按用户反馈修订《项目进度》Excel：三大未达标问题如实下调。"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

PATH = r"C:\Users\J\WorkBuddy\2026-08-17-08-58-49\淘宝巡检系统_项目进度_20260821.xlsx"
wb = openpyxl.load_workbook(PATH)

BODY_FONT = Font(name="微软雅黑", size=10)
HEAD_FONT = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
STATUS_FILL = {
    "已完成": PatternFill("solid", fgColor="C6EFCE"),
    "进行中": PatternFill("solid", fgColor="FFEB9C"),
    "存在严重问题": PatternFill("solid", fgColor="FFC7CE"),
}
STATUS_FONT = {
    "已完成": Font(name="微软雅黑", size=10, color="006100"),
    "进行中": Font(name="微软雅黑", size=10, color="9C6500"),
    "存在严重问题": Font(name="微软雅黑", size=10, bold=True, color="9C0006"),
}
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

def pct_fill(pct):
    p = int(str(pct).strip("%"))
    color = "C6EFCE" if p >= 100 else "FFEB9C" if p >= 80 else "FFC7CE"
    return PatternFill("solid", fgColor=color)

def find_row(ws, col, text):
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=col).value
        if v and str(v).strip().startswith(text):
            return r
    return None

def set_status(ws, r, status, note=None):
    sc = ws.cell(row=r, column=4)
    sc.value = status
    sc.fill = STATUS_FILL.get(status, PatternFill())
    sc.font = STATUS_FONT.get(status, BODY_FONT)
    sc.alignment = CENTER
    if note is not None:
        nc = ws.cell(row=r, column=5)
        base = str(nc.value or "")
        if note not in base:
            nc.value = base.rstrip("；;") + "；" + note if base else note
        nc.alignment = LEFT

# ===================== Sheet1: 模块进度总览 =====================
ws = wb["模块进度总览"]

# (1) 图片自动抓取 -> 严重不准确
r = find_row(ws, 2, "图片自动抓取")
if r:
    ws.cell(row=r, column=3, value="40%").fill = pct_fill("40%")
    ws.cell(row=r, column=3).alignment = CENTER
    ws.cell(row=r, column=3).font = Font(name="微软雅黑", size=10, bold=True)
    set_status(ws, r, "存在严重问题",
               "主图/SKU/详情抓取分类或召回结果不准确，巡店图片抓取质量严重不达标——准确率问题优先于功能完整性，需重点修复")

# (2) 违规看板 / (3) 左右对比视图 -> 面板需优化
for label in ("违规看板", "左右对比视图"):
    r = find_row(ws, 2, label)
    if r:
        ws.cell(row=r, column=3, value="70%").fill = pct_fill("70%")
        ws.cell(row=r, column=3).alignment = CENTER
        ws.cell(row=r, column=3).font = Font(name="微软雅黑", size=10, bold=True)
        set_status(ws, r, "进行中",
                   "功能可用但操作面板/交互体验需大幅优化，未达体验验收标准")

# (4) 新增「网站稳定性 / 服务常驻」行（插在汇总行之前）
sum_row = None
for rr in range(ws.max_row, 4, -1):
    if ws.cell(row=rr, column=2).value and "加权完成度" in str(ws.cell(row=rr, column=2).value):
        sum_row = rr
        break
if sum_row:
    ws.insert_rows(sum_row)  # 新空白行 = sum_row，旧汇总下移到 sum_row+1
    nr = sum_row
    ws.cell(row=nr, column=1, value=14).alignment = CENTER
    ws.cell(row=nr, column=2, value="网站稳定性 / 服务常驻").alignment = LEFT
    pc = ws.cell(row=nr, column=3, value="50%")
    pc.fill = pct_fill("50%"); pc.alignment = CENTER
    pc.font = Font(name="微软雅黑", size=10, bold=True)
    set_status(ws, nr, "进行中",
               "服务存在不稳定（崩溃/重复实例抢库 readonly、依赖平台保活），尚未达到「所有人可稳定访问」")
    for c in (1, 2, 3, 4):
        ws.cell(row=nr, column=c).border = BORDER
        if ws.cell(row=nr, column=c).font is None:
            ws.cell(row=nr, column=c).font = BODY_FONT
    # 汇总行下移后重算
    new_sum = nr + 1
    vals = []
    for rr in range(5, nr + 1):
        v = ws.cell(row=rr, column=3).value
        if v:
            try: vals.append(int(str(v).strip("%")))
            except Exception: pass
    avg = round(sum(vals) / len(vals)) if vals else 0
    avgcell = ws.cell(row=new_sum, column=3, value=f"{avg}%")
    avgcell.fill = PatternFill("solid", fgColor=("C6EFCE" if avg >= 100 else "FFEB9C" if avg >= 80 else "FFC7CE"))
    avgcell.font = Font(name="微软雅黑", size=10, bold=True)
    avgcell.alignment = CENTER
    lc = ws.cell(row=new_sum, column=2)
    lc.value = "整体（14 模块）加权完成度"
    sc = ws.cell(row=new_sum, column=4, value="核心三项未达标")
    sc.fill = STATUS_FILL["进行中"]; sc.font = STATUS_FONT["进行中"]; sc.alignment = CENTER

# ===================== Sheet2: 图片抓取专项 =====================
ws2 = wb["图片抓取专项"]
r = find_row(ws2, 2, "Playwright 抓取脚本是否已调通")
if r:
    ws2.cell(row=r, column=2, value="严重问题（不准确）").fill = STATUS_FILL["存在严重问题"]
    ws2.cell(row=r, column=2).font = STATUS_FONT["存在严重问题"]
    ws2.cell(row=r, column=2).alignment = CENTER
    ws2.cell(row=r, column=3,
             value="脚本可运行但当前抓取结果（主图/SKU/详情分类与召回）不准确，属严重质量问题；优先级由「功能联通」转为「准确率修复」，需先解决抓取准确性再谈功能完整性")

# ===================== Sheet4: 风险与待办 =====================
ws4 = wb["风险与待办"]
# 清空原数据行（5 起），重写为新优先级
for rr in range(5, ws4.max_row + 1):
    for c in range(1, 5):
        ws4.cell(row=rr, column=c, value=None)
PRIO_FILL = {"P0": "FFC7CE", "P1": "FFEB9C", "P2": "DDEBF7"}
rows4 = [
    ("P0", "图片抓取准确性修复", "巡店图片抓取严重不准确，直接影响巡检结论可信度，是当前最大阻塞项。",
     "暂停新功能，先定位分类/召回不准根因（选择器、Cookie 时效、页面结构变化），建立准确率验收基准后再回归。"),
    ("P0", "网站稳定性", "服务崩溃 / 重复实例抢库 readonly、依赖平台保活，导致「所有人可稳定访问」未达成，存在线上不可用风险。",
     "治理常驻进程（单实例锁 / 防重复启动）、增加崩溃自恢复与平台保活，做稳定性压测。"),
    ("P0", "操作面板 / UI 优化", "功能可用但交互体验未达标准，影响使用效率与验收。",
     "梳理操作面板高频路径，做交互 / 布局重构，设定体验验收标准后再判定完成。"),
    ("P1", "推送至内网 Git 远程", "当前仅本地仓库，交接即「拷文件夹」，易丢版本。", "提供 Git 地址后 remote add + push。"),
    ("P1", "公网部署取得 URL", "CloudStudio 仅支持静态，本项目需 Python 托管。", "用已备好的 Railway/Render 配置部署，或自有服务器。"),
    ("P2", "滑块验证 / IP 限流", "反爬极端情形未自动处理，靠过期重扫兜底。", "观察触发频次，如高频再引入代理 / 打码服务。"),
]
r = 5
for prio, item, risk, act in rows4:
    ws4.cell(row=r, column=1, value=prio).fill = PatternFill("solid", fgColor=PRIO_FILL[prio])
    ws4.cell(row=r, column=1).font = Font(name="微软雅黑", size=10, bold=True)
    ws4.cell(row=r, column=1).alignment = CENTER
    ws4.cell(row=r, column=2, value=item).alignment = LEFT
    ws4.cell(row=r, column=3, value=risk).alignment = LEFT
    ws4.cell(row=r, column=4, value=act).alignment = LEFT
    for c in range(1, 5):
        ws4.cell(row=r, column=c).border = BORDER
        if ws4.cell(row=r, column=c).font is None:
            ws4.cell(row=r, column=c).font = BODY_FONT
    r += 1

wb.save(PATH)
print("SAVED:", PATH)
print("Sheet1 rows:", ws.max_row, "| Sheet4 rows:", ws4.max_row)
