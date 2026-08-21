import openpyxl

path = r'C:\Users\J\WorkBuddy\2026-08-17-08-58-49\淘宝巡检系统_项目进度_20260821.xlsx'
wb = openpyxl.load_workbook(path)

# ---- 模块进度总览：第10行 AI 视觉比对 如实下调 ----
ws = wb['模块进度总览']
ws.cell(row=10, column=3, value='35%')
ws.cell(row=10, column=4, value='存在严重问题')
ws.cell(row=10, column=5,
        value='算法代码已写完（主图轮播/SKU色卡/详情iframe 三类分抓比对+差异标注），'
              '但本质是传统像素相似度（SSIM+pHash），并非真正的 AI 语义合规；'
              '代理店与官旗店页面模板不同导致大面积误报/漏报，且结果被「抓取不准」严重污染，'
              '比对结论不可信——需重构为语义/模型级视觉合规检测。')

# ---- 重算整体加权（第5~18行） ----
total, n = 0, 0
for r in range(5, 19):
    v = ws.cell(row=r, column=3).value
    if isinstance(v, str) and v.endswith('%'):
        total += float(v[:-1]); n += 1
avg = round(total / n) if n else 0
ws.cell(row=19, column=3, value=f'{avg}%')
ws.cell(row=19, column=4, value='核心四项未达标')
ws.cell(row=19, column=5,
        value='抓取40% / AI视觉比对35% / 网站稳定性50% / 操作面板70% 均未达验收；'
              '比对结论不可信与稳定性风险最高。')

# ---- 风险与待办：在第5行(P0 图片抓取)后插入新 P0 ----
ws4 = wb['风险与待办']
ws4.insert_rows(6)
ws4.cell(row=6, column=1, value='P0')
ws4.cell(row=6, column=2, value='AI 视觉对比引擎重构（像素哈希 → 语义/模型级合规检测）')
ws4.cell(row=6, column=3,
         value='当前 SSIM/pHash 像素比对在「代理店页面 ≠ 官旗店页面模板」场景下大面积误报/漏报；'
               '且强烈依赖抓取质量，抓取不准时比对结论完全不可信。')
ws4.cell(row=6, column=4,
         value='暂停依赖现有比对结论；重构为语义级检测（接入多模态视觉模型或鲁棒特征比对），'
               '先建立准确率验收基准再回归。')

wb.save(path)
print('OK 整体加权 =', avg, '%')
